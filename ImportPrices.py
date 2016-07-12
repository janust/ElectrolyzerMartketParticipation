# -*- coding: utf-8 -*-
"""
Created on Fri Jan 15 16:22:08 2016

@author: Janus
"""

# Import the electricity data from Energinet.dk source Excel file.
# End product is a dictionary priceWDate

def importPrices(priceColumn):
    
    from datetime import datetime, date, time, timedelta
    from openpyxl import load_workbook
    
    tMax = 8760*6+24
    #dMax = int(tMax/24)
    
    tSet = list(range(1,tMax+1))
    
    wb = load_workbook(filename = 'MarketData20102015.xlsx')
    ws = wb['MarketData']
    
    
    ############### Need to  make a separate import of the data.
    #  Define function or class that can extract a subset of the data and make e.g.
    #  price scenarios basesd of the subset and average price forecast.
    
    
    # Import the price from Excel
    price = {}
    for t in tSet:
        price[t] = ws[priceColumn+str(3+t)].value
    
    # Import the corresponding date from Excel
    datoTid = {}
    for t in tSet:
        datoTid[t] = ws['A'+str(3+t)].value
        # For some reason, energinet.dk changes randomly btw datetime and str number format
        # This formats str to datetime
        if type(datoTid[t]) == type(str()):
            datoTid[t] = date(int(datoTid[t][6:11]),int(datoTid[t][0:2]),
            int(datoTid[t][3:5]))
        datoTid[t] = datetime.combine(datoTid[t],time(ws['B'+str(3+t)].value-1))
        
    priceWDate = {}
    for t in tSet:
        priceWDate[datoTid[t]] = price[t]
    
    i = 0    
    for datTid in list(priceWDate.keys()):
        if type(priceWDate[datTid]) == type(None):
            # The energinet.dk data-set is empty two days a year because of summer/winther time.
            # For now, we'll just fill in the value from the hour before.
            priceWDate[datTid] = priceWDate[datTid-timedelta(0,0,0,0,0,1)]
            i += 1
    print('Number of type \'None\' price enteries in data-set: '+str(i))
    
    #dkkEUR = 7.5
        
    #tarEnergiFyn = (16.23+8.20)*10/dkkEUR # Energi Fyn City, elpristavlen.dk
    #tarPSO = 26.10*10/dkkEUR # Energi Fyn City, elpristavlen.dk
    #tarTAX = 0.4*10/dkkEUR # Development of Solar Energy Plants, Assignment 2
    
    #priceTaxWDate = {}
    #for datTid in list(priceWDate.keys()):
    #    priceTaxWDate[datTid] = priceWDate[datTid] + tarEnergiFyn + tarPSO + tarTAX
    
    return priceWDate
 
def AsUpDownNone(priceWDate,balPriceWDate):
    AsUpDownNoneDict = {}  
    for k in priceWDate.keys():
        if priceWDate[k] < balPriceWDate[k]:
            AsUpDownNoneDict[k] = 'Up'
        if priceWDate[k] > balPriceWDate[k]:
            AsUpDownNoneDict[k] = 'Down'
        if priceWDate[k] == balPriceWDate[k]:
            AsUpDownNoneDict[k] = 'None'
    return AsUpDownNoneDict
    
def AsDownUpBinary(AsUpDownNoneDict,Direction):
    yAs = {}
    for k in AsUpDownNoneDict.keys():
        if AsUpDownNoneDict[k] == Direction:
            yAs[k] = 1
        else:
            yAs[k] = 0
    return yAs
    

# Import price, bal price and analyze for up and down reg   
priceWDate = importPrices('C')
balPriceWDate = importPrices('F')
balPriceUp = balPriceWDate
balPriceDown = balPriceWDate
AsUpDownNoneDict = AsUpDownNone(priceWDate,balPriceWDate)
yAsUp = AsDownUpBinary(AsUpDownNoneDict,'Up')
yAsDown = AsDownUpBinary(AsUpDownNoneDict,'Down')

# Analyze amount of up and down regulation in the market
#up = 0
#down = 0
#for k in AsUpDownNoneDict.keys():
#    if AsUpDownNoneDict[k] == 'Up':
#        up += 1
#    if AsUpDownNoneDict[k] == 'Down':
#        down += 1
#print(up/len(AsUpDownNoneDict))
#print(down/len(AsUpDownNoneDict))





    